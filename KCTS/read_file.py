from openpyxl import load_workbook
import re

# 加载 Excel 文件
workbook = load_workbook(filename="test.xlsx")
# 查看 OU->MU 表单
ou_to_mu = workbook['OU->MU']

# 初始化协议字典
ou_protocol = {}

# 遍历所有合并的单元格
for merged_range in ou_to_mu.merged_cells.ranges:
    # 合并单元格的开始和终止坐标
    start_idx_str, end_idx_str = merged_range.coord.split(':')
    # 合并单元格的起始和终止单元格
    start_cell, end_cell = ou_to_mu[start_idx_str], ou_to_mu[end_idx_str]
    if start_cell.value == '开关量':
        # 遍历开关量的字节序号
        byte_col = start_cell.column + 1
        for byte_row in range(start_cell.row, end_cell.row + 1):
            # 合并的字节序号如果为空就跳过
            if ou_to_mu.cell(row=byte_row, column=byte_col).value:
                # 获取开关量的字节序号
                if isinstance(ou_to_mu.cell(row=byte_row, column=byte_col).value, str):
                    byte_num = int(re.findall(r'\d', ou_to_mu.cell(row=byte_row, column=byte_col).value)[0])
                else:
                    byte_num = ou_to_mu.cell(row=byte_row, column=byte_col).value
                # 更新协议的字节号
                ou_protocol.update({byte_num:{}})
                # 将字节序号的行列转换为坐标形式 A1
                byte_coord = ou_to_mu.cell(row=byte_row, column=byte_col).coordinate
                # 检查字节序号是合并的单元格
                is_merged = None
                for is_merged in ou_to_mu.merged_cells.ranges:
                    if byte_coord in is_merged:
                        # 如果是合并的单元格就更新坐标
                        byte_coord = is_merged.coord
                        # 获取字节序号合并单元格的开始和终止坐标
                        start_byte_idx, end_byte_idx = byte_coord.split(':')
                        # 字节序号起始和终止的单元格
                        start_byte_cell, end_byte_cell = ou_to_mu[start_byte_idx], ou_to_mu[end_byte_idx]
                        # 遍历对应字节的bit位
                        bit_col = start_byte_cell.column + 1
                        for bit_row in range(start_byte_cell.row, end_byte_cell.row + 1):
                            # 获取bit位的索引
                            if isinstance(ou_to_mu.cell(row=bit_col, column=bit_col).value, str):
                                bit_index = int(re.findall(r'\d', ou_to_mu.cell(row=bit_row, column=bit_col).value)[0])
                            else:
                                bit_index = ou_to_mu.cell(row=bit_row, column=bit_col).value
                            # 获取开关量的协议内容(规定bit内容的列数在bit位右边一列)
                            funciton_col = bit_col + 1
                            bit_function = ou_to_mu.cell(row=bit_row, column=funciton_col).value
                            # 更新协议字典
                            ou_protocol.get(byte_num).update({bit_index: bit_function})
                    # 如果开关量的字节序号不是合并单元格
                    # else:
                    #     bit_row, bit_col, function_col = byte_row, byte_col + 1, byte_col + 2
                    #     print(f'现在的行是{bit_row}现在的列是{byte_col}')
                    #     # 获取bit位索引
                    #     if isinstance(ou_to_mu.cell(row=bit_row, column=bit_col).value, str):
                    #         bit_index = int(re.findall(r'\d', ou_to_mu.cell(row=bit_row, column=bit_col).value)[0])
                    #     else:
                    #         bit_index = ou_to_mu.cell(row=bit_row, column=bit_col).value
                    #     bit_function = ou_to_mu.cell(row=bit_col, column=funciton_col).value
                    #     # 更新协议字典
                    #     ou_protocol.get(byte_num).update({bit_index: bit_function})

    elif start_cell.value == '模拟量':
        # 遍历模拟量的字节序号
        byte_col = start_cell.column + 1
        for byte_row in range(start_cell.row, end_cell.row + 1):
            # 获取模拟量的字节序号
            if isinstance(ou_to_mu.cell(row=byte_row, column=byte_col).value, str):
                byte_num = int(re.findall(r'\d', ou_to_mu.cell(row=byte_row, column=byte_col).value)[0])
            else:
                byte_num = ou_to_mu.cell(row=byte_row, column=byte_col).value
            # 获取模拟量的协议内容(规定字节内容的列数为字节序号 + 2)
            funciton_col = byte_col + 2
            byte_function = ou_to_mu.cell(row=byte_row, column=funciton_col).value
            # 更新协议字典
            if byte_function == '预留':
                continue
            else:
                ou_protocol.update({byte_num: byte_function})

print(ou_protocol)