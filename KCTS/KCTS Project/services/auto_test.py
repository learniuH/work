from PyQt5.QtWidgets import QFileDialog, QTableWidget, QPushButton, QLineEdit, QLabel, QComboBox, QTableWidgetItem, \
    QHeaderView

from widget.constant import ConstantText

from openpyxl import load_workbook
import pandas as pd
import numpy as np


class AutoTest:
    """ 自动化测试类 """

    def __init__(self, widget: dict):
        """ 界面子控件初始化 """
        self.tableWidget_test_case: QTableWidget        =   widget['tableWidget_test_case']
        self.comboBox_case_sheet_names: QComboBox       =   widget['comboBox_case_sheet_names']
        self.pushButton_insert_case: QPushButton        =   widget['pushButton_insert_case']
        self.pushButton_start_test: QPushButton         =   widget['pushButton_start_test']
        self.pushButton_stop_test: QPushButton          =   widget['pushButton_stop_test']
        self.pushButton_previous_page: QPushButton      =   widget['pushButton_previous_page']
        self.pushButton_next_page: QPushButton          =   widget['pushButton_next_page']
        self.lineEdit_current_page: QLineEdit           =   widget['lineEdit_current_page']
        self.label_total_pages: QLabel                  =   widget['label_total_pages']

        self.file_path: str             =   None    # 测试用例路径
        self.target_df: pd.DataFrame    =   None    # 解析后的 Excel 表格数据
        self.total_pages: int           =   None    # TableWidget 的总页数
        self.page_index: int            =   None    # TableWidget 当前页索引

        self.slot_connect()

    def slot_connect(self):
        """ 自动化测试界面子控件槽链接 """
        self.pushButton_insert_case.clicked.connect(lambda: self.open_dialog())
        # 信号重载(传递当前文本 str)
        # self.comboBox_case_sheet_names.currentIndexChanged[str].connect(
        #     lambda current_text: self.excel_case_parse(file_path=file_path, sheet_name=current_text))
        self.comboBox_case_sheet_names.currentIndexChanged[str].connect(self.excel_case_parse)
        # self.pushButton_start_test.clicked.connect(lambda: )
        # self.pushButton_stop_test.clicked.connect(lambda: )
        # self.pushButton_previous_page.clicked.connect()
        # self.pushButton_next_page.clicked.connect()
        # self.lineEdit_current_page.textChanged.connect()

    def open_dialog(self):
        """ 点击导入用例, 更新 comboBox内容 """
        self.file_path, _ = QFileDialog.getOpenFileName(
            parent= None,
            caption= "选择测试用例",
            filter= "ExcelFile (*.xlsx *.xls)",
        )
        if self.file_path:
            # 获取用例 Excel 的表单列表
            sheet_names = self.get_sheet_names(self.file_path)
            # 将表单名更新到 comboBox
            self.update_comboBox_items(sheet_names)

    def get_sheet_names(self, file_path: str) -> list:
        """
        根据测试用例路径, 解析Excel表单列表
        :param file_path:  测试用例Excel的路径
        :return: 测试用例Excel的表单列表
        """
        excel_file = pd.ExcelFile(file_path)
        sheet_names: list = excel_file.sheet_names
        return sheet_names

    def update_comboBox_items(self, sheet_names: list):
        """ 根据表单名列表更新 comboBox 的 items """
        self.comboBox_case_sheet_names.blockSignals(True)
        # 清除所有 items
        self.comboBox_case_sheet_names.clear()
        self.comboBox_case_sheet_names.addItems(sheet_names)
        self.comboBox_case_sheet_names.setCurrentIndex(-1)

        self.comboBox_case_sheet_names.blockSignals(False)

    def get_merged_cells(self, sheet_name: str) -> list:
        """
        取合并单元格信息
        :param sheet_name: 要解析的Excel表单的名称
        :return: 返回该表单内所有合并单元格的最小/最大行号, 最小/最大列号
                 (min_row, min_col, max_row, max_col), 行号/列号按照Excel格式, 从1开始
        """
        xls_wb = load_workbook(self.file_path)
        xls_sheet = xls_wb[sheet_name]
        return [
            (merged_range.min_row, merged_range.min_col,
             merged_range.max_row, merged_range.max_col)
            for merged_range in xls_sheet.merged_cells.ranges
        ]


    def excel_case_parse(self, sheet_name: str):
        """ 根据表单名称, 解析Excel内容 """
        df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None)
        # 将"用例编号"所在行设置为 columns
        all_rows, all_cols = np.where(df == '用例编号')
        # 选取行号最小的作为新的 columns, 并删除该行
        df.columns = df.iloc[all_rows[0]]
        self.target_df = df.iloc[all_rows[0] + 1:, :]

        # 页面初始化, 展示首页内容
        self.total_pages = (len(self.target_df) // ConstantText.ROWS_PER_PAGES +
                            (1 if len(self.target_df) % ConstantText.ROWS_PER_PAGES else 0))
        self.page_index = 0
        self.show_page(self.page_index)


    def show_page(self, page_index: int):
        """ 在 TableWidget 上显示当前页面数据 """
        # 计算当前页面的开始行索引和终止行索引
        start_row = page_index * ConstantText.ROWS_PER_PAGES
        end_row = min((page_index + 1) * ConstantText.ROWS_PER_PAGES, len(self.target_df))

        # 设置表格的行数
        self.tableWidget_test_case.setRowCount(end_row - start_row)

        # 填充表格内容
        for row in range(start_row, end_row):
            for col, value_header in enumerate(ConstantText.COLUMNS_TO_INSERT):
                item = QTableWidgetItem(str(self.target_df.loc[row + 1, value_header]))
                self.tableWidget_test_case.setItem(row, col, item)

            # 设置大于一行的行高, 实现自动换行
            self.tableWidget_test_case.setRowHeight(row, 100)

        # 根据内容调整自适应行高
        self.tableWidget_test_case.resizeRowsToContents()